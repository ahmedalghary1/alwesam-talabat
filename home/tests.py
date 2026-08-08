import os
import shutil
import tempfile
from io import BytesIO
from unittest.mock import patch

from PIL import Image
from django.contrib import admin
from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import RequestFactory, SimpleTestCase, TestCase, override_settings
from django.urls import reverse
from django.utils.datastructures import MultiValueDict
from openpyxl import load_workbook

from orders.models import Order
from products.models import (
    Category, Product, ProductImages, ProductSize, ProductVariant, Size,
    ProductSizeImage, VariantAttribute, VariantAttributeValue, VariantImage,
    VariantSize, VariantSizeImage,
)
from products.admin import ProductSizeInline, VariantSizeInline
from support.models import CustomerMessage, MessageReply

from .admin_views import _excel_safe_text, _valid_model_ids, _variant_sizes_from_post


class ValidModelIdsTests(SimpleTestCase):
    def test_ignores_blank_and_invalid_values(self):
        self.assertEqual(
            _valid_model_ids(['', '  ', 'invalid', None, '7']),
            [7],
        )

    def test_accepts_only_positive_integer_ids(self):
        self.assertEqual(
            _valid_model_ids(['1', '0', '-2', '3']),
            [1, 3],
        )

    def test_variant_sizes_keep_the_quantity_for_each_checked_size(self):
        request = RequestFactory().post('/', {
            'variant_7_size_ids[]': ['2', '9'],
            'variant_7_size_pcs_2': '12',
            'variant_7_size_pcs_9': '30',
        })

        self.assertEqual(_variant_sizes_from_post(request, '7'), [(2, 12), (9, 30)])

    def test_excel_text_cannot_be_interpreted_as_a_formula(self):
        self.assertEqual(_excel_safe_text('=2+2'), "'=2+2")
        self.assertEqual(_excel_safe_text('+201000000000'), "'+201000000000")
        self.assertEqual(_excel_safe_text('بيانات عربية'), 'بيانات عربية')


def _image_file(name='test.png', color='red'):
    output = BytesIO()
    Image.new('RGB', (20, 20), color).save(output, format='PNG')
    return SimpleUploadedFile(name, output.getvalue(), content_type='image/png')


class ProductAdminFlowTests(TestCase):
    def setUp(self):
        self.media_root = tempfile.mkdtemp()
        self.override = override_settings(MEDIA_ROOT=self.media_root)
        self.override.enable()
        self.addCleanup(self.override.disable)
        self.addCleanup(shutil.rmtree, self.media_root, True)

        user_model = get_user_model()
        self.staff = user_model.objects.create_user(
            email='admin@example.com', username='admin', password='password',
            phone='01000000000', address='test', is_staff=True, is_active=True,
        )
        self.client.force_login(self.staff)

    def test_edit_uses_stable_variant_key_for_availability_and_images(self):
        product = Product.objects.create(name='منتج', image=_image_file())
        removed = ProductVariant.objects.create(product=product, name='قديم')
        kept = ProductVariant.objects.create(product=product, name='مستمر', is_available=False)

        response = self.client.post(
            f'/admin-panel/products/{product.pk}/edit/',
            {
                'name': product.name,
                'pcs_carton': '24',
                'order': '0',
                'is_available': 'on',
                'variant_form_key[]': ['7'],
                'variant_id[]': [str(kept.pk)],
                'variant_name[]': ['مستمر'],
                'variant_code[]': [''],
                'variant_pcs_carton[]': ['24'],
                'variant_available[]': ['7'],
                'variant_color[]': [''],
                'variant_length_label[]': [''],
                'variant_7_new_images[]': _image_file('variant.png', 'blue'),
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertFalse(ProductVariant.objects.filter(pk=removed.pk).exists())
        kept.refresh_from_db()
        self.assertTrue(kept.is_available)
        self.assertEqual(VariantImage.objects.filter(variant=kept).count(), 1)

    def test_add_product_saves_the_quantity_for_each_variant_size(self):
        first_size = Size.objects.create(name='مقاس أول')
        second_size = Size.objects.create(name='مقاس ثان')

        response = self.client.post(
            '/admin-panel/products/add/',
            {
                'name': 'منتج بكميات مقاسات',
                'image': _image_file('sized-product.png'),
                'pcs_carton': '24',
                'is_available': 'on',
                'variant_form_key[]': ['0'],
                'variant_name[]': ['النمط الأول'],
                'variant_code[]': [''],
                'variant_pcs_carton[]': ['30'],
                'variant_available[]': ['0'],
                'variant_color[]': [''],
                'variant_length_label[]': [''],
                'variant_0_size_ids[]': [str(first_size.pk), str(second_size.pk)],
                'variant_0_size_pcs[]': ['12', '60'],
            },
        )

        self.assertEqual(response.status_code, 302)
        variant = ProductVariant.objects.get(product__name='منتج بكميات مقاسات')
        self.assertTrue(VariantSize.objects.filter(
            variant=variant, size=first_size, pcs_carton=12,
        ).exists())
        self.assertTrue(VariantSize.objects.filter(
            variant=variant, size=second_size, pcs_carton=60,
        ).exists())

    def test_add_product_saves_images_for_direct_and_variant_sizes(self):
        direct_size = Size.objects.create(name='Direct image size')
        second_direct_size = Size.objects.create(name='Second direct image size')
        variant_size_name = Size.objects.create(name='Variant image size')
        second_variant_size = Size.objects.create(name='Second variant image size')

        response = self.client.post(
            reverse('admin_app:admin_product_add'),
            {
                'name': 'Product with size images',
                'image': _image_file('main-size-image.png'),
                'pcs_carton': '24',
                'product_size_ids[]': [str(direct_size.pk), str(second_direct_size.pk)],
                'product_size_pcs[]': ['48', '60'],
                f'product_size_{direct_size.pk}_images[]': _image_file('direct-size.png'),
                f'product_size_{second_direct_size.pk}_images[]': _image_file('second-direct-size.png', 'blue'),
                'variant_form_key[]': ['0'],
                'variant_name[]': ['Image variant'],
                'variant_code[]': [''],
                'variant_pcs_carton[]': ['30'],
                'variant_available[]': ['0'],
                'variant_length_label[]': [''],
                'variant_0_size_ids[]': [str(variant_size_name.pk), str(second_variant_size.pk)],
                'variant_0_size_pcs[]': ['72', '84'],
                f'variant_0_size_{variant_size_name.pk}_images[]': _image_file('variant-size.png'),
                f'variant_0_size_{second_variant_size.pk}_images[]': _image_file('second-variant-size.png', 'green'),
            },
        )

        self.assertEqual(response.status_code, 302)
        product = Product.objects.get(name='Product with size images')
        first_direct = product.size_prices.get(size=direct_size)
        second_direct = product.size_prices.get(size=second_direct_size)
        self.assertEqual(first_direct.pcs_carton, 48)
        self.assertEqual(second_direct.pcs_carton, 60)
        self.assertEqual(first_direct.images.count(), 1)
        self.assertEqual(second_direct.images.count(), 1)
        variant = product.variants.get()
        first_variant_size = variant.size_prices.get(size=variant_size_name)
        second_variant = variant.size_prices.get(size=second_variant_size)
        self.assertEqual(first_variant_size.pcs_carton, 72)
        self.assertEqual(second_variant.pcs_carton, 84)
        self.assertEqual(first_variant_size.images.count(), 1)
        self.assertEqual(second_variant.images.count(), 1)
        self.assertNotEqual(
            first_direct.images.get().image.name,
            second_direct.images.get().image.name,
        )
        self.assertNotEqual(
            first_variant_size.images.get().image.name,
            second_variant.images.get().image.name,
        )

    def test_editing_variant_size_quantity_preserves_its_images(self):
        product = Product.objects.create(name='Persistent size gallery', image=_image_file())
        variant = ProductVariant.objects.create(product=product, name='Persistent variant')
        size = Size.objects.create(name='Persistent size')
        variant_size = VariantSize.objects.create(
            variant=variant, size=size, pcs_carton=24
        )
        size_image = VariantSizeImage.objects.create(
            variant_size=variant_size, image=_image_file('persistent-size.png')
        )

        response = self.client.post(
            reverse('admin_app:admin_product_edit', args=[product.pk]),
            {
                'name': product.name,
                'pcs_carton': '24',
                'order': '0',
                'variants_form_initialized': '1',
                'variant_form_key[]': ['0'],
                'variant_id[]': [str(variant.pk)],
                'variant_name[]': [variant.name],
                'variant_code[]': [''],
                'variant_pcs_carton[]': ['24'],
                'variant_available[]': ['0'],
                'variant_length_label[]': [''],
                'variant_0_size_ids[]': [str(size.pk)],
                f'variant_0_size_pcs_{size.pk}': '60',
            },
        )

        self.assertEqual(response.status_code, 302)
        variant_size.refresh_from_db()
        self.assertEqual(variant_size.pcs_carton, 60)
        self.assertTrue(VariantSizeImage.objects.filter(pk=size_image.pk).exists())

    def test_edit_adds_different_images_to_each_existing_size(self):
        product = Product.objects.create(name='Editable size galleries', image=_image_file())
        direct_one = ProductSize.objects.create(
            product=product,
            size=Size.objects.create(name='Editable direct one'),
            pcs_carton=24,
        )
        direct_two = ProductSize.objects.create(
            product=product,
            size=Size.objects.create(name='Editable direct two'),
            pcs_carton=36,
        )
        variant = ProductVariant.objects.create(product=product, name='Editable variant')
        variant_one = VariantSize.objects.create(
            variant=variant,
            size=Size.objects.create(name='Editable variant one'),
            pcs_carton=48,
        )
        variant_two = VariantSize.objects.create(
            variant=variant,
            size=Size.objects.create(name='Editable variant two'),
            pcs_carton=60,
        )

        response = self.client.post(
            reverse('admin_app:admin_product_edit', args=[product.pk]),
            {
                'name': product.name,
                'pcs_carton': '24',
                'order': '0',
                'product_size_id[]': [str(direct_one.pk), str(direct_two.pk)],
                'product_size_pcs[]': ['24', '36'],
                f'product_size_{direct_one.size_id}_new_images[]': _image_file('edit-direct-one.png', 'red'),
                f'product_size_{direct_two.size_id}_new_images[]': _image_file('edit-direct-two.png', 'blue'),
                'variants_form_initialized': '1',
                'variant_form_key[]': ['0'],
                'variant_id[]': [str(variant.pk)],
                'variant_name[]': [variant.name],
                'variant_code[]': [''],
                'variant_pcs_carton[]': ['24'],
                'variant_available[]': ['0'],
                'variant_length_label[]': [''],
                'variant_0_size_ids[]': [str(variant_one.size_id), str(variant_two.size_id)],
                f'variant_0_size_pcs_{variant_one.size_id}': '48',
                f'variant_0_size_pcs_{variant_two.size_id}': '60',
                f'variant_0_size_{variant_one.size_id}_new_images[]': _image_file('edit-variant-one.png', 'green'),
                f'variant_0_size_{variant_two.size_id}_new_images[]': _image_file('edit-variant-two.png', 'yellow'),
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(direct_one.images.count(), 1)
        self.assertEqual(direct_two.images.count(), 1)
        self.assertEqual(variant_one.images.count(), 1)
        self.assertEqual(variant_two.images.count(), 1)
        self.assertNotEqual(direct_one.images.get().image.name, direct_two.images.get().image.name)
        self.assertNotEqual(variant_one.images.get().image.name, variant_two.images.get().image.name)

    def test_custom_and_django_admin_show_size_image_upload_controls(self):
        product = Product.objects.create(name='Visible size uploads', image=_image_file())
        size = Size.objects.create(name='Visible direct size')
        ProductSize.objects.create(product=product, size=size, pcs_carton=24)
        variant = ProductVariant.objects.create(product=product, name='Visible variant')
        VariantSize.objects.create(
            variant=variant,
            size=Size.objects.create(name='Visible variant size'),
            pcs_carton=36,
        )
        self.staff.is_superuser = True
        self.staff.save(update_fields=['is_superuser'])

        responses = [
            self.client.get(reverse('admin_app:admin_product_add')),
            self.client.get(reverse('admin_app:admin_product_edit', args=[product.pk])),
            self.client.get(reverse('admin:products_product_change', args=[product.pk])),
            self.client.get(reverse('admin:products_productvariant_change', args=[variant.pk])),
        ]

        self.assertTrue(all(response.status_code == 200 for response in responses))
        for response in responses:
            self.assertContains(response, 'صور خاصة بهذا المقاس')

    def test_django_admin_inlines_save_multiple_images_per_size(self):
        product = Product.objects.create(name='Django admin size uploads', image=_image_file())
        product_size = ProductSize.objects.create(
            product=product,
            size=Size.objects.create(name='Django direct size'),
            pcs_carton=24,
        )
        variant = ProductVariant.objects.create(product=product, name='Django variant')
        variant_size = VariantSize.objects.create(
            variant=variant,
            size=Size.objects.create(name='Django variant size'),
            pcs_carton=36,
        )
        self.staff.is_superuser = True
        self.staff.save(update_fields=['is_superuser'])
        request = RequestFactory().post('/')
        request.user = self.staff

        cases = [
            (ProductSizeInline, product, product_size, 'direct'),
            (VariantSizeInline, variant, variant_size, 'variant'),
        ]
        for inline_class, parent, size_relation, file_prefix in cases:
            inline = inline_class(type(parent), admin.site)
            formset_class = inline.get_formset(request, parent)
            prefix = 'size_prices'
            data = {
                f'{prefix}-TOTAL_FORMS': '1',
                f'{prefix}-INITIAL_FORMS': '1',
                f'{prefix}-MIN_NUM_FORMS': '0',
                f'{prefix}-MAX_NUM_FORMS': '1000',
                f'{prefix}-0-id': str(size_relation.pk),
                f'{prefix}-0-size': str(size_relation.size_id),
                f'{prefix}-0-pcs_carton': str(size_relation.pcs_carton),
            }
            files = MultiValueDict({
                f'{prefix}-0-size_images': [
                    _image_file(f'{file_prefix}-one.png', 'red'),
                    _image_file(f'{file_prefix}-two.png', 'blue'),
                ]
            })
            formset = formset_class(
                data=data,
                files=files,
                instance=parent,
                prefix=prefix,
            )
            self.assertTrue(formset.is_valid(), formset.errors)
            model_admin = admin.site._registry[type(parent)]
            model_admin.save_formset(request, None, formset, change=True)

            self.assertEqual(size_relation.images.count(), 2)

    def test_repeated_arabic_product_names_get_unique_slugs(self):
        first = Product.objects.create(name='منتج مكرر', image=_image_file('one.png'))
        second = Product.objects.create(name='منتج مكرر', image=_image_file('two.png'))

        self.assertEqual(first.slug, 'منتج-مكرر')
        self.assertEqual(second.slug, 'منتج-مكرر-2')
        self.assertTrue(first.image.name.endswith('.webp'))

    def test_edit_keeps_a_new_direct_size(self):
        product = Product.objects.create(name='منتج مقاسات', image=_image_file())
        size = Size.objects.create(name='كبير')

        response = self.client.post(
            f'/admin-panel/products/{product.pk}/edit/',
            {
                'name': product.name,
                'pcs_carton': '24',
                'order': '0',
                'new_product_size_id[]': [str(size.pk)],
                'new_product_size_pcs[]': ['48'],
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertTrue(ProductSize.objects.filter(
            product=product, size=size, pcs_carton=48,
        ).exists())

    def test_edit_page_renders_product_data(self):
        product = Product.objects.create(name='منتج "آمن"', image=_image_file())
        ProductVariant.objects.create(product=product, name='<نمط>')

        response = self.client.get(f'/admin-panel/products/{product.pk}/edit/')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="variants-data"')
        self.assertNotContains(response, '<نمط>')

    def test_edit_page_includes_all_variants_and_attribute_types(self):
        product = Product.objects.create(name='منتج متعدد الأنماط', image=_image_file())
        material = VariantAttribute.objects.create(name='الخامة')
        cotton = VariantAttributeValue.objects.create(attribute=material, value='قطن')
        variants = [
            ProductVariant.objects.create(product=product, name=f'نمط {number}')
            for number in range(1, 4)
        ]
        variants[-1].attributes.add(cotton)

        response = self.client.get(f'/admin-panel/products/{product.pk}/edit/')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context['variants_data']), 3)
        self.assertEqual(
            response.context['variants_data'][-1]['attribute_ids'], [cotton.pk]
        )
        self.assertEqual(
            response.context['attribute_groups_data'][0]['values'][0]['name'],
            'قطن',
        )
        self.assertContains(response, 'id="attribute-groups-data"')

    def test_edit_saves_and_clears_all_variant_attribute_types(self):
        product = Product.objects.create(name='منتج الخصائص', image=_image_file())
        material = VariantAttribute.objects.create(name='الخامة')
        cotton = VariantAttributeValue.objects.create(attribute=material, value='قطن')
        model = VariantAttribute.objects.create(name='الموديل')
        modern = VariantAttributeValue.objects.create(attribute=model, value='مودرن')
        variant = ProductVariant.objects.create(product=product, name='نمط الخصائص')
        base_data = {
            'name': product.name,
            'pcs_carton': '24',
            'order': '0',
            'variant_form_key[]': ['4'],
            'variant_id[]': [str(variant.pk)],
            'variant_name[]': [variant.name],
            'variant_code[]': [''],
            'variant_pcs_carton[]': ['24'],
            'variant_available[]': ['4'],
            'variant_length_label[]': [''],
        }

        response = self.client.post(
            f'/admin-panel/products/{product.pk}/edit/',
            {
                **base_data,
                'variant_4_attribute_ids[]': ['', str(cotton.pk), str(modern.pk)],
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertSetEqual(
            set(variant.attributes.values_list('pk', flat=True)),
            {cotton.pk, modern.pk},
        )

        response = self.client.post(
            f'/admin-panel/products/{product.pk}/edit/',
            {**base_data, 'variant_4_attribute_ids[]': ['']},
        )

        self.assertEqual(response.status_code, 302)
        self.assertFalse(variant.attributes.exists())

    def test_edit_deletes_selected_product_and_variant_images_only(self):
        product = Product.objects.create(name='منتج الصور', image=_image_file())
        other_product = Product.objects.create(name='منتج آخر', image=_image_file('other-main.png'))
        product_image = ProductImages.objects.create(
            product=product, image=_image_file('delete-product.png')
        )
        other_image = ProductImages.objects.create(
            product=other_product, image=_image_file('keep-product.png')
        )
        variant = ProductVariant.objects.create(product=product, name='نمط الصور')
        variant_image = VariantImage.objects.create(
            variant=variant, image=_image_file('delete-variant.png')
        )
        product_image_path = product_image.image.path
        variant_image_path = variant_image.image.path

        with self.captureOnCommitCallbacks(execute=True):
            response = self.client.post(
                f'/admin-panel/products/{product.pk}/edit/',
                {
                    'name': product.name,
                    'pcs_carton': '24',
                    'order': '0',
                    'variant_form_key[]': ['0'],
                    'variant_id[]': [str(variant.pk)],
                    'variant_name[]': [variant.name],
                    'variant_code[]': [''],
                    'variant_pcs_carton[]': ['24'],
                    'variant_available[]': ['0'],
                    'variant_length_label[]': [''],
                    'delete_product_images[]': [str(product_image.pk)],
                    'delete_variant_images[]': [str(variant_image.pk)],
                },
            )

        self.assertEqual(response.status_code, 302)
        self.assertFalse(ProductImages.objects.filter(pk=product_image.pk).exists())
        self.assertFalse(VariantImage.objects.filter(pk=variant_image.pk).exists())
        self.assertTrue(ProductImages.objects.filter(pk=other_image.pk).exists())
        self.assertFalse(os.path.exists(product_image_path))
        self.assertFalse(os.path.exists(variant_image_path))

    def test_edit_without_initialized_javascript_keeps_existing_variants(self):
        product = Product.objects.create(name='منتج آمن', image=_image_file())
        variant = ProductVariant.objects.create(product=product, name='نمط محفوظ')

        response = self.client.post(
            reverse('admin_app:admin_product_edit', args=[product.pk]),
            {'name': product.name, 'pcs_carton': '24', 'order': '0'},
        )

        self.assertEqual(response.status_code, 302)
        self.assertTrue(ProductVariant.objects.filter(pk=variant.pk).exists())

    def test_add_product_saves_product_and_variant_order(self):
        response = self.client.post(
            reverse('admin_app:admin_product_add'),
            {
                'name': 'منتج مرتب',
                'image': _image_file('ordered.png'),
                'pcs_carton': '24',
                'order': '8',
                'variant_form_key[]': ['0'],
                'variant_name[]': ['نمط مرتب'],
                'variant_code[]': [''],
                'variant_pcs_carton[]': ['24'],
                'variant_order[]': ['6'],
                'variant_available[]': ['0'],
                'variant_length_label[]': [''],
            },
        )

        self.assertEqual(response.status_code, 302)
        product = Product.objects.get(name='منتج مرتب')
        self.assertEqual(product.order, 8)
        self.assertEqual(product.variants.get().order, 6)


class AdminPanelReliabilityTests(TestCase):
    def setUp(self):
        self.media_root = tempfile.mkdtemp()
        self.override = override_settings(MEDIA_ROOT=self.media_root)
        self.override.enable()
        self.addCleanup(self.override.disable)
        self.addCleanup(shutil.rmtree, self.media_root, True)

        user_model = get_user_model()
        self.staff = user_model.objects.create_user(
            email='panel-admin@example.com',
            username='panel-admin',
            password='password',
            phone='01000000011',
            address='الإدارة',
            is_staff=True,
            is_active=True,
        )
        self.customer = user_model.objects.create_user(
            email='panel-customer@example.com',
            username='عميل اللوحة',
            password='password',
            phone='01000000012',
            address='عنوان العميل',
            is_active=False,
        )
        self.client.force_login(self.staff)
        self.category = Category.objects.create(
            name='قسم الاختبار', image=_image_file('category.png')
        )
        self.product = Product.objects.create(
            name='منتج لوحة الإدارة',
            category=self.category,
            image=_image_file('panel-product.png'),
        )
        self.order = Order.objects.create(
            user=self.customer,
            phone_number=self.customer.phone,
            address=self.customer.address,
            status='confirmed',
        )
        self.message = CustomerMessage.objects.create(
            user=self.customer,
            message='رسالة اختبار لوحة الإدارة',
        )

    def test_main_admin_pages_render_with_working_controls(self):
        page_urls = [
            reverse('admin_app:dashboard'),
            reverse('admin_app:admin_products'),
            reverse('admin_app:admin_categories'),
            reverse('admin_app:admin_orders'),
            reverse('admin_app:admin_order_detail', args=[self.order.pk]),
            reverse('admin_app:pending_users'),
            reverse('admin_app:all_users'),
            reverse('admin_support:messages_list'),
            reverse('admin_support:conversation_detail', args=[self.message.pk]),
        ]

        responses = [self.client.get(url) for url in page_urls]

        self.assertTrue(all(response.status_code == 200 for response in responses))
        dashboard, products, _, orders, _, pending, _, messages_page, conversation = responses
        self.assertContains(dashboard, self.order.get_status_display())
        self.assertContains(products, 'name="search"')
        self.assertContains(products, 'name="category"')
        self.assertContains(orders, self.order.get_status_display())
        self.assertContains(pending, 'onclick="showUserDetails(this)"')
        self.assertContains(messages_page, "filterMessages('all', this)")
        self.assertNotContains(messages_page, 'event.target')
        self.assertContains(conversation, f'محادثة مع {self.customer.username}')
        self.assertContains(conversation, 'admin-layout')

    def test_product_search_and_category_filter_work_together(self):
        Product.objects.create(name='منتج خارج البحث', image=_image_file('outside.png'))

        response = self.client.get(reverse('admin_app:admin_products'), {
            'search': 'لوحة الإدارة',
            'category': self.category.slug,
        })

        self.assertContains(response, self.product.name)
        self.assertNotContains(response, 'منتج خارج البحث')

    def test_category_validation_order_and_rename_work(self):
        response = self.client.post(reverse('admin_app:admin_category_add'), {
            'name': 'قسم بلا صورة',
            'order': '3',
        })
        self.assertEqual(response.status_code, 200)
        self.assertFalse(Category.objects.filter(name='قسم بلا صورة').exists())

        response = self.client.post(
            reverse('admin_app:admin_category_edit', args=[self.category.pk]),
            {'name': 'قسم بعد التعديل', 'description': 'وصف', 'order': '9'},
        )
        self.assertEqual(response.status_code, 302)
        self.category.refresh_from_db()
        self.assertEqual(self.category.name, 'قسم بعد التعديل')
        self.assertEqual(self.category.order, 9)
        self.assertEqual(self.category.slug, 'قسم-بعد-التعديل')

    @patch('utils.email_tasks.send_order_status_email_task.delay')
    def test_order_status_update_redirects_and_confirms_success(self, delay):
        response = self.client.post(
            reverse('admin_app:admin_order_detail', args=[self.order.pk]),
            {'status': 'delivered'},
            follow=True,
        )

        self.order.refresh_from_db()
        self.assertEqual(self.order.status, 'delivered')
        self.assertContains(response, 'تم تحديث حالة الطلب بنجاح')
        delay.assert_called_once()

    @patch('utils.email_tasks.send_order_status_email_task.delay')
    def test_order_status_without_email_does_not_schedule_email(self, delay):
        self.customer.email = None
        self.customer.save(update_fields=['email'])

        response = self.client.post(
            reverse('admin_app:admin_order_detail', args=[self.order.pk]),
            {'status': 'delivered'},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        delay.assert_not_called()

    @patch('utils.email_tasks.send_activation_email_task.delay')
    def test_approve_user_without_email_does_not_schedule_email(self, delay):
        customer = get_user_model().objects.create_user(
            username='approval-without-email',
            email=None,
            phone='01000000019',
            address='Cairo',
            password='password',
            is_active=False,
        )

        response = self.client.post(
            reverse('admin_app:approve_user', args=[customer.pk]),
            follow=True,
        )

        customer.refresh_from_db()
        self.assertTrue(customer.is_active)
        self.assertEqual(response.status_code, 200)
        delay.assert_not_called()

    def test_mutating_admin_endpoints_reject_get_requests(self):
        urls = [
            reverse('admin_app:admin_product_delete', args=[self.product.pk]),
            reverse('admin_app:admin_category_delete', args=[self.category.pk]),
            reverse('admin_app:approve_user', args=[self.customer.pk]),
            reverse('admin_app:reject_user', args=[self.customer.pk]),
            reverse('admin_app:toggle_user_status', args=[self.customer.pk]),
            reverse('admin_support:send_reply', args=[self.message.pk]),
            reverse('admin_support:mark_as_read', args=[self.message.pk]),
            reverse('admin_support:delete_message', args=[self.message.pk]),
        ]

        responses = [self.client.get(url) for url in urls]

        self.assertTrue(all(response.status_code == 405 for response in responses))
        self.assertTrue(Product.objects.filter(pk=self.product.pk).exists())
        self.assertTrue(Category.objects.filter(pk=self.category.pk).exists())
        self.assertTrue(CustomerMessage.objects.filter(pk=self.message.pk).exists())

    def test_support_ajax_reply_returns_and_saves_reply(self):
        response = self.client.post(
            reverse('admin_support:send_reply', args=[self.message.pk]),
            {'reply': 'رد من الإدارة'},
            HTTP_X_REQUESTED_WITH='XMLHttpRequest',
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['success'])
        self.assertTrue(MessageReply.objects.filter(
            customer_message=self.message,
            reply='رد من الإدارة',
        ).exists())

    def test_image_delete_buttons_use_direct_post_endpoints(self):
        product = Product.objects.create(name='منتج الحذف المباشر', image=_image_file())
        product_image = ProductImages.objects.create(
            product=product, image=_image_file('direct-product.png')
        )
        variant = ProductVariant.objects.create(product=product, name='نمط مباشر')
        variant_image = VariantImage.objects.create(
            variant=variant, image=_image_file('direct-variant.png')
        )
        direct_size = ProductSize.objects.create(
            product=product,
            size=Size.objects.create(name='Direct image deletion size'),
            pcs_carton=24,
        )
        direct_size_image = ProductSizeImage.objects.create(
            product_size=direct_size,
            image=_image_file('direct-size-delete.png'),
        )
        variant_size = VariantSize.objects.create(
            variant=variant,
            size=Size.objects.create(name='Variant image deletion size'),
            pcs_carton=24,
        )
        variant_size_image = VariantSizeImage.objects.create(
            variant_size=variant_size,
            image=_image_file('variant-size-delete.png'),
        )
        product_image_path = product_image.image.path
        variant_image_path = variant_image.image.path
        direct_size_image_path = direct_size_image.image.path
        variant_size_image_path = variant_size_image.image.path
        product_delete_url = reverse(
            'admin_app:admin_product_image_delete',
            args=[product.pk, product_image.pk],
        )
        variant_delete_url = reverse(
            'admin_app:admin_variant_image_delete',
            args=[product.pk, variant_image.pk],
        )
        direct_size_delete_url = reverse(
            'admin_app:admin_product_size_image_delete',
            args=[product.pk, direct_size_image.pk],
        )
        variant_size_delete_url = reverse(
            'admin_app:admin_variant_size_image_delete',
            args=[product.pk, variant_size_image.pk],
        )

        page_response = self.client.get(
            reverse('admin_app:admin_product_edit', args=[product.pk])
        )
        self.assertContains(page_response, product_delete_url)
        self.assertContains(page_response, variant_delete_url)
        self.assertContains(page_response, direct_size_delete_url)
        self.assertContains(page_response, variant_size_delete_url)
        self.assertContains(page_response, 'deleteImageImmediately')

        with self.captureOnCommitCallbacks(execute=True):
            product_response = self.client.post(product_delete_url)
            variant_response = self.client.post(variant_delete_url)
            direct_size_response = self.client.post(direct_size_delete_url)
            variant_size_response = self.client.post(variant_size_delete_url)

        self.assertEqual(product_response.status_code, 200)
        self.assertTrue(product_response.json()['deleted'])
        self.assertEqual(variant_response.status_code, 200)
        self.assertTrue(variant_response.json()['deleted'])
        self.assertEqual(direct_size_response.status_code, 200)
        self.assertTrue(direct_size_response.json()['deleted'])
        self.assertEqual(variant_size_response.status_code, 200)
        self.assertTrue(variant_size_response.json()['deleted'])
        self.assertFalse(ProductImages.objects.filter(pk=product_image.pk).exists())
        self.assertFalse(VariantImage.objects.filter(pk=variant_image.pk).exists())
        self.assertFalse(ProductSizeImage.objects.filter(pk=direct_size_image.pk).exists())
        self.assertFalse(VariantSizeImage.objects.filter(pk=variant_size_image.pk).exists())
        self.assertFalse(os.path.exists(product_image_path))
        self.assertFalse(os.path.exists(variant_image_path))
        self.assertFalse(os.path.exists(direct_size_image_path))
        self.assertFalse(os.path.exists(variant_size_image_path))

    def test_direct_image_delete_cannot_delete_another_products_image(self):
        product = Product.objects.create(name='المنتج المطلوب', image=_image_file())
        other_product = Product.objects.create(
            name='منتج آخر', image=_image_file('other-direct-main.png')
        )
        other_image = ProductImages.objects.create(
            product=other_product, image=_image_file('other-direct-extra.png')
        )

        response = self.client.post(reverse(
            'admin_app:admin_product_image_delete',
            args=[product.pk, other_image.pk],
        ))

        self.assertEqual(response.status_code, 404)
        self.assertTrue(ProductImages.objects.filter(pk=other_image.pk).exists())


class UserExcelExportTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.staff = user_model.objects.create_user(
            email='admin@example.com',
            username='admin',
            password='password',
            phone='01000000000',
            address='القاهرة',
            is_staff=True,
            is_active=True,
        )
        self.user = user_model.objects.create_user(
            email='customer@example.com',
            username='عميل-الوسام',
            password='customer-password',
            phone='01012345678',
            address='القاهرة، مدينة نصر',
            first_name='محمد',
            last_name='علي',
            is_active=True,
        )
        self.user.profile.bio = 'عميل جملة'
        self.user.profile.save()
        Order.objects.create(
            user=self.user,
            phone_number=self.user.phone,
            address=self.user.address,
        )
        CustomerMessage.objects.create(user=self.user, message='رسالة اختبار')
        self.export_url = reverse('admin_app:export_users_excel')

    def test_only_staff_can_export_users(self):
        response = self.client.get(self.export_url)

        self.assertEqual(response.status_code, 302)
        self.assertIn('/admin/login/', response.url)

    def test_users_page_shows_the_excel_export_button(self):
        self.client.force_login(self.staff)

        response = self.client.get(reverse('admin_app:all_users'))

        self.assertContains(response, self.export_url)
        self.assertContains(response, 'استخراج جميع المستخدمين إلى Excel')

    def test_export_is_a_complete_arabic_excel_workbook(self):
        self.client.force_login(self.staff)
        self.staff.profile.delete()

        response = self.client.get(self.export_url)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('.xlsx', response['Content-Disposition'])
        workbook = load_workbook(BytesIO(response.content), data_only=False)
        worksheet = workbook['المستخدمون']

        self.assertTrue(worksheet.sheet_view.rightToLeft)
        self.assertEqual(worksheet.freeze_panes, 'A5')
        headers = [cell.value for cell in worksheet[4]]
        self.assertIn('رقم الهاتف', headers)
        self.assertIn('النبذة الشخصية', headers)
        self.assertIn('عدد الطلبات', headers)
        self.assertNotIn('كلمة المرور', headers)

        rows = list(worksheet.iter_rows(min_row=5, values_only=True))
        customer_row = next(row for row in rows if row[0] == self.user.pk)
        self.assertEqual(customer_row[1], 'عميل-الوسام')
        self.assertEqual(customer_row[4], 'محمد علي')
        self.assertEqual(customer_row[6], '01012345678')
        self.assertEqual(customer_row[7], 'القاهرة، مدينة نصر')
        self.assertEqual(customer_row[8], 'عميل جملة')
        self.assertEqual(customer_row[18], 1)
        self.assertEqual(customer_row[20], 1)

        exported_ids = {row[0] for row in rows}
        self.assertEqual(exported_ids, {self.staff.pk, self.user.pk})
