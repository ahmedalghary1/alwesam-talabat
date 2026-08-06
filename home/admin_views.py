from io import BytesIO
from urllib.parse import quote

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from django.db import transaction
from django.db.models import Q, Count, Max
from django.db.models.deletion import ProtectedError
from django.http import HttpResponse, JsonResponse
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST
from products.models import (
    Product, Category, ProductImages, ProductVariant,
    VariantAttributeValue, VariantAttribute, Size,
    ProductSize, VariantSize, VariantImage
)
from orders.models import Order, OrderItem


def _valid_model_ids(values):
    """Return only positive integer IDs from untrusted form values."""
    valid_ids = []
    for value in values:
        try:
            model_id = int(value)
        except (TypeError, ValueError):
            continue
        if model_id > 0:
            valid_ids.append(model_id)
    return valid_ids


def _positive_int(value, default=None):
    """Parse a positive integer from an untrusted form value."""
    try:
        value = int(value)
    except (TypeError, ValueError):
        if default is not None:
            return default
        raise ValueError('يجب إدخال رقم صحيح موجب')
    if value < 1:
        raise ValueError('يجب أن تكون الكمية 1 على الأقل')
    return value


def _variant_sizes_from_post(request, form_key):
    """Read variant sizes without relying on fragile parallel checkbox lists."""
    size_ids = request.POST.getlist(f'variant_{form_key}_size_ids[]')
    legacy_quantities = request.POST.getlist(f'variant_{form_key}_size_pcs[]')
    result = []
    seen = set()
    for index, size_id in enumerate(size_ids):
        size_id = _positive_int(size_id)
        if size_id in seen:
            continue
        quantity = request.POST.get(f'variant_{form_key}_size_pcs_{size_id}')
        if quantity is None and index < len(legacy_quantities):
            quantity = legacy_quantities[index]
        result.append((size_id, _positive_int(quantity, 24)))
        seen.add(size_id)
    return result


def _attribute_groups_data():
    """Serialize every variant attribute group for the custom admin forms."""
    groups = VariantAttribute.objects.prefetch_related('values').order_by('name')
    return [
        {
            'id': group.id,
            'name': group.name,
            'values': [
                {'id': value.id, 'name': value.value}
                for value in sorted(
                    group.values.all(), key=lambda item: item.value.casefold()
                )
            ],
        }
        for group in groups
    ]


def _set_variant_attributes_from_post(request, variant, form_key, colors=None, color_id=None):
    """Save all submitted attributes while accepting the legacy color field."""
    field_name = f'variant_{form_key}_attribute_ids[]'
    if field_name in request.POST:
        attribute_ids = _valid_model_ids(request.POST.getlist(field_name))
        variant.attributes.set(VariantAttributeValue.objects.filter(id__in=attribute_ids))
        return

    old_colors = variant.attributes.filter(attribute__name__in=["لون", "Color"])
    if old_colors.exists():
        variant.attributes.remove(*old_colors)
    if color_id and colors is not None:
        try:
            variant.attributes.add(colors.get(id=color_id))
        except VariantAttributeValue.DoesNotExist:
            pass


@staff_member_required
def admin_dashboard(request):
    """Admin dashboard with statistics and recent orders"""
    total_products = Product.objects.count()
    total_categories = Category.objects.count()
    total_orders = Order.objects.count()
    pending_orders = Order.objects.filter(status='pending').count()

    recent_orders = Order.objects.all().order_by('-created_at')[:10]

    context = {
        'total_products': total_products,
        'total_categories': total_categories,
        'total_orders': total_orders,
        'pending_orders': pending_orders,
        'recent_orders': recent_orders,
    }
    return render(request, 'admin/dashboard.html', context)


@staff_member_required
def admin_products(request):
    """Product management with search and filter"""
    search_query = request.GET.get('search', '')
    category_filter = request.GET.get('category', '')

    products = Product.objects.all()

    if search_query:
        products = products.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query)
        )

    if category_filter:
        products = products.filter(category__slug=category_filter)

    categories = Category.objects.all()

    return render(request, 'admin/products.html', {
        'products': products,
        'categories': categories,
        'search_query': search_query,
        'category_filter': category_filter,
    })


@staff_member_required
def admin_product_add(request):
    """Add a new product with direct sizes and variant sizes (with quantities)"""
    categories = Category.objects.all()
    colors = VariantAttributeValue.objects.filter(
        attribute__name__in=["لون", "Color"]
    ).all()
    sizes = Size.objects.all()
    attribute_groups_data = _attribute_groups_data()

    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        pcs_carton = request.POST.get('pcs_carton', 1)
        category_id = request.POST.get('category')
        image = request.FILES.get('image')

        try:
            if not name or not name.strip():
                raise ValueError('اسم المنتج مطلوب')
            if not image:
                raise ValueError('صورة المنتج الرئيسية مطلوبة')
            with transaction.atomic():
                # Create product
                product = Product.objects.create(
                    name=name.strip(),
                    slug='',
                    description=description,
                    pcs_carton=_positive_int(pcs_carton, 1),
                    category_id=category_id if category_id else None,
                    image=image,
                    is_available='is_available' in request.POST
                )

                # Additional images for product
                additional_images = request.FILES.getlist('additional_images[]')
                for idx, img in enumerate(additional_images):
                    ProductImages.objects.create(product=product, image=img, order=idx)

            # ========== Direct product sizes (ProductSize) ==========
                product_size_ids = request.POST.getlist('product_size_ids[]')
                product_size_pcs = request.POST.getlist('product_size_pcs[]')
                for size_id, pcs in zip(product_size_ids, product_size_pcs):
                    if size_id and pcs:
                        ProductSize.objects.update_or_create(
                            product=product,
                            size_id=_positive_int(size_id),
                            defaults={'pcs_carton': _positive_int(pcs)}
                        )

            # ========== Variants ==========
                variant_names = request.POST.getlist('variant_name[]')
                variant_codes = request.POST.getlist('variant_code[]')
                variant_pcs = request.POST.getlist('variant_pcs_carton[]')
                variant_available = request.POST.getlist('variant_available[]')
                variant_color_ids = request.POST.getlist('variant_color[]')
                variant_length_labels = request.POST.getlist('variant_length_label[]')
                variant_keys = request.POST.getlist('variant_form_key[]')

                for i in range(len(variant_names)):
                    if not variant_names[i].strip():
                        continue
                    form_key = variant_keys[i] if i < len(variant_keys) else str(i)
                    # Create variant
                    variant = ProductVariant.objects.create(
                        product=product,
                        name=variant_names[i].strip(),
                        code=variant_codes[i] if i < len(variant_codes) and variant_codes[i] else None,
                        pcs_carton=_positive_int(variant_pcs[i], 24) if i < len(variant_pcs) else 24,
                        is_available=form_key in variant_available,
                        length_label=(variant_length_labels[i] or None) if i < len(variant_length_labels) else None,
                    )

                    # Add every selected attribute (color, material, model, etc.).
                    color_id = (
                        variant_color_ids[i]
                        if i < len(variant_color_ids) and variant_color_ids[i]
                        else None
                    )
                    _set_variant_attributes_from_post(
                        request, variant, form_key, colors=colors, color_id=color_id
                    )

                    # ========== Variant sizes with quantities (VariantSize) ==========
                    for size_id, pcs in _variant_sizes_from_post(request, form_key):
                        VariantSize.objects.create(variant=variant, size_id=size_id, pcs_carton=pcs)

                    # ========== Variant images ==========
                    variant_images_key = f'variant_{form_key}_images[]'
                    variant_images = request.FILES.getlist(variant_images_key)
                    for idx, img in enumerate(variant_images):
                        VariantImage.objects.create(
                            variant=variant,
                            image=img,
                            order=idx
                        )

            messages.success(request, f'تم إضافة المنتج "{product.name}" بنجاح')
            return redirect('admin_app:admin_products')

        except Exception as e:
            messages.error(request, f'حدث خطأ: {str(e)}')
            import traceback
            traceback.print_exc()

    return render(request, 'admin/product_add.html', {
        'categories': categories,
        'colors': colors,
        'sizes': sizes,
        'attribute_groups_data': attribute_groups_data,
    })


@staff_member_required
@transaction.atomic
def admin_product_edit(request, product_id):
    """Edit product with direct sizes and variant sizes (with quantities)"""
    product = get_object_or_404(Product, id=product_id)
    categories = Category.objects.all()
    colors = VariantAttributeValue.objects.filter(
        attribute__name__in=["لون", "Color"]
    ).select_related('attribute').all()
    sizes = Size.objects.all().order_by('order')
    attribute_groups_data = _attribute_groups_data()

    # Prepare variants with their size data (through VariantSize)
    variants = product.variants.prefetch_related(
        'attributes__attribute',
        'images'
    ).all().order_by('order')

    for variant in variants:
        variant_attributes = list(variant.attributes.all())
        color_attr = next(
            (
                attribute
                for attribute in variant_attributes
                if attribute.attribute.name in ["لون", "Color"]
            ),
            None,
        )
        variant.selected_color_id = color_attr.id if color_attr else None
        variant.selected_attribute_ids = [attribute.id for attribute in variant_attributes]
        # Get size data with pcs_carton from VariantSize
        variant.size_data = list(variant.size_prices.select_related('size').values(
            'id', 'size_id', 'size__name', 'pcs_carton'
        ))

    # Build JSON data for variants (for the template)
    variants_data = []
    for variant in variants:
        variant_dict = {
            'id': variant.id,
            'name': variant.name,
            'code': variant.code,
            'pcs_carton': variant.pcs_carton,
            'is_available': variant.is_available,
            'color_id': variant.selected_color_id,
            'attribute_ids': variant.selected_attribute_ids,
            'length_label': variant.length_label,
            'size_data': list(variant.size_prices.select_related('size').values(
                'size_id', 'pcs_carton'
            )),
            'images': [
                {
                    'id': img.id,
                    'url': img.image.url,
                    'delete_url': reverse(
                        'admin_app:admin_variant_image_delete',
                        args=[product.id, img.id],
                    ),
                }
                for img in variant.images.all()
            ]
        }
        variants_data.append(variant_dict)

    # Direct product sizes with quantities
    product_size_data = list(product.size_prices.select_related('size').values(
        'id', 'size_id', 'size__name', 'pcs_carton'
    ))

    product_images = product.additional_images.all().order_by('order')

    if request.method == 'POST':
        try:
            # ========== Update product basic info ==========
            old_name = product.name
            product.name = request.POST.get('name', product.name).strip()
            if not product.name:
                raise ValueError('اسم المنتج مطلوب')
            product.description = request.POST.get('description', product.description)
            product.pcs_carton = _positive_int(request.POST.get('pcs_carton'), product.pcs_carton)
            product.is_available = 'is_available' in request.POST
            product.order = max(int(request.POST.get('order') or 0), 0)

            category_id = request.POST.get('category')
            product.category_id = category_id if category_id else None

            if 'image' in request.FILES:
                product.image = request.FILES['image']

            if old_name != product.name or not product.slug:
                product.slug = ''

            product.save()

            # ========== Product additional images ==========
            # Delete selected images
            delete_product_images = _valid_model_ids(
                request.POST.getlist('delete_product_images[]')
            )
            if delete_product_images:
                ProductImages.objects.filter(id__in=delete_product_images, product=product).delete()

            # Add new images
            new_product_images = request.FILES.getlist('new_product_images[]')
            if new_product_images:
                from django.db import models as db_models
                max_order = product.additional_images.aggregate(db_models.Max('order'))['order__max'] or 0
                for idx, img in enumerate(new_product_images):
                    ProductImages.objects.create(
                        product=product,
                        image=img,
                        order=max_order + idx + 1
                    )

            # Update order of existing images
            image_orders = request.POST.getlist('image_order[]')
            image_ids = request.POST.getlist('image_id[]')
            for img_id, order in zip(image_ids, image_orders):
                if img_id and order != '':
                    ProductImages.objects.filter(id=img_id, product=product).update(order=max(int(order), 0))

            # ========== Direct product sizes (ProductSize) ==========
            existing_ps_ids = request.POST.getlist('product_size_id[]')
            existing_ps_pcs = request.POST.getlist('product_size_pcs[]')
            new_ps_size_ids = request.POST.getlist('new_product_size_id[]')
            new_ps_pcs = request.POST.getlist('new_product_size_pcs[]')

            # Update existing
            for ps_id, pcs in zip(existing_ps_ids, existing_ps_pcs):
                if ps_id and pcs:
                    ProductSize.objects.filter(id=ps_id, product=product).update(pcs_carton=_positive_int(pcs))

            # Delete removed existing rows before creating new rows. Otherwise the
            # newly created rows are absent from existing_ps_ids and get deleted.
            kept_ps_ids = _valid_model_ids(existing_ps_ids)
            product.size_prices.exclude(id__in=kept_ps_ids).delete()

            # Add new
            for size_id, pcs in zip(new_ps_size_ids, new_ps_pcs):
                if size_id and pcs:
                    ProductSize.objects.update_or_create(
                        product=product,
                        size_id=_positive_int(size_id),
                        defaults={'pcs_carton': _positive_int(pcs)}
                    )

            # ========== Variants ==========
            variant_ids = request.POST.getlist('variant_id[]')
            variant_names = request.POST.getlist('variant_name[]')
            variant_codes = request.POST.getlist('variant_code[]')
            variant_pcs = request.POST.getlist('variant_pcs_carton[]')
            variant_order = request.POST.getlist('variant_order[]')
            variant_available = request.POST.getlist('variant_available[]')
            variant_color_ids = request.POST.getlist('variant_color[]')
            variant_length_labels = request.POST.getlist('variant_length_label[]')
            variant_keys = request.POST.getlist('variant_form_key[]')

            updated_variant_ids = []
            updated_variants = []

            for i in range(len(variant_names)):
                if variant_names[i].strip():
                    form_key = variant_keys[i] if i < len(variant_keys) else str(i)
                    variant_data = {
                        'product': product,
                        'name': variant_names[i].strip(),
                        'code': variant_codes[i] if i < len(variant_codes) and variant_codes[i] else None,
                        'pcs_carton': _positive_int(variant_pcs[i], 24) if i < len(variant_pcs) else 24,
                        'is_available': form_key in variant_available,
                        'length_label': (variant_length_labels[i] or None) if i < len(variant_length_labels) else None,
                    }
                    if i < len(variant_order) and variant_order[i]:
                        variant_data['order'] = max(int(variant_order[i]), 0)

                    color_id = variant_color_ids[i] if i < len(variant_color_ids) and variant_color_ids[i] else None

                    if i < len(variant_ids) and variant_ids[i]:
                        # Update existing variant
                        try:
                            variant = ProductVariant.objects.get(id=variant_ids[i], product=product)
                            for key, value in variant_data.items():
                                setattr(variant, key, value)
                            variant.save()

                            updated_variant_ids.append(variant.id)
                            updated_variants.append((form_key, variant))
                        except ProductVariant.DoesNotExist:
                            # If ID invalid, create new
                            variant = ProductVariant.objects.create(**variant_data)
                            updated_variant_ids.append(variant.id)
                            updated_variants.append((form_key, variant))
                    else:
                        # Create new variant
                        variant = ProductVariant.objects.create(**variant_data)
                        updated_variant_ids.append(variant.id)
                        updated_variants.append((form_key, variant))

                    _set_variant_attributes_from_post(
                        request, variant, form_key, colors=colors, color_id=color_id
                    )

                    # ========== Variant sizes (VariantSize) ==========
                    # Delete all existing sizes for this variant and recreate
                    variant.size_prices.all().delete()
                    for size_id, pcs in _variant_sizes_from_post(request, form_key):
                        VariantSize.objects.create(variant=variant, size_id=size_id, pcs_carton=pcs)

            # ========== Delete variants that were removed ==========
            deleted_count = product.variants.exclude(id__in=updated_variant_ids).delete()
            if deleted_count[0] > 0:
                messages.info(request, f'تم حذف {deleted_count[0]} نمط/أنماط')

            # ========== Variant images ==========
            # Delete selected images
            delete_variant_images = _valid_model_ids(
                request.POST.getlist('delete_variant_images[]')
            )
            if delete_variant_images:
                VariantImage.objects.filter(
                    id__in=delete_variant_images,
                    variant__product=product,
                ).delete()

            # Add new images for each variant
            for form_key, variant in updated_variants:
                new_images_key = f'variant_{form_key}_new_images[]'
                new_images = request.FILES.getlist(new_images_key)
                if new_images:
                    max_order = variant.images.aggregate(Max('order'))['order__max'] or 0
                    for img_idx, img in enumerate(new_images):
                        VariantImage.objects.create(
                            variant=variant,
                            image=img,
                            order=max_order + img_idx + 1
                        )

            messages.success(request, f'تم تحديث المنتج "{product.name}" بنجاح')
            return redirect('admin_app:admin_products')

        except Exception as e:
            transaction.set_rollback(True)
            messages.error(request, f'حدث خطأ أثناء تحديث المنتج: {str(e)}')
            return redirect('admin_app:admin_product_edit', product_id=product.pk)

    context = {
        'product': product,
        'categories': categories,
        'colors': colors,
        'sizes': sizes,
        'variants': variants,
        'product_images': product_images,
        'product_size_data': product_size_data,
        'variants_data': variants_data,
        'sizes_data': list(sizes.values('id', 'name')),
        'colors_data': [
            {'id': color.id, 'name': str(color)} for color in colors
        ],
        'attribute_groups_data': attribute_groups_data,
        'total_variants': variants.count(),
        'total_images': product_images.count(),
    }
    return render(request, 'admin/product_edit.html', context)


@staff_member_required
@require_POST
def admin_product_image_delete(request, product_id, image_id):
    """Delete one additional product image directly from the custom admin."""
    product = get_object_or_404(Product, id=product_id)
    image = get_object_or_404(ProductImages, id=image_id, product=product)
    image.delete()
    return JsonResponse({'deleted': True, 'image_id': image_id})


@staff_member_required
@require_POST
def admin_variant_image_delete(request, product_id, image_id):
    """Delete one variant image directly from the custom admin."""
    image = get_object_or_404(
        VariantImage,
        id=image_id,
        variant__product_id=product_id,
    )
    image.delete()
    return JsonResponse({'deleted': True, 'image_id': image_id})


@staff_member_required
def admin_product_delete(request, product_id):
    """Delete product"""
    if request.method == 'POST':
        product = get_object_or_404(Product, id=product_id)
        product_name = product.name
        try:
            product.delete()
            messages.success(request, f'تم حذف المنتج {product_name}')
        except ProtectedError:
            product.is_available = False
            product.save(update_fields=['is_available', 'updated_at'])
            messages.warning(
                request,
                f'لا يمكن حذف المنتج {product_name} لأنه موجود في طلبات سابقة. '
                'تم إخفاؤه من المتجر بدلًا من ذلك.',
            )
    return redirect('admin_app:admin_products')


@staff_member_required
def admin_orders(request):
    """Order management with search and filter"""
    status_filter = request.GET.get('status', '')
    search_query = request.GET.get('search', '').strip()

    orders = Order.objects.all().select_related('user').prefetch_related(
        'items__product',
        'items__variant'
    ).order_by('-created_at')

    if search_query:
        orders = orders.filter(
            Q(id__icontains=search_query) |
            Q(user__username__icontains=search_query)
        ).distinct()

    if status_filter:
        orders = orders.filter(status=status_filter)

    return render(request, 'admin/orders.html', {
        'orders': orders,
        'status_filter': status_filter,
        'search_query': search_query,
    })


@staff_member_required
def admin_order_detail(request, order_id):
    """Order details in admin panel, including variant information"""
    order = get_object_or_404(Order, id=order_id)
    # Prefetch items with related product and variant
    order_items = order.items.select_related('product', 'variant').all()

    if request.method == 'POST':
        new_status = request.POST.get('status')
        if new_status in dict(Order.STATUS_CHOICES).keys():
            order.status = new_status
            order.save()

            try:
                from utils.email_tasks import send_order_status_email_task
                send_order_status_email_task.delay(order.id, new_status, order.user.email)
            except Exception as e:
                messages.success(request, f'تم تحديث حالة الطلب (فشل جدولة إرسال البريد الإلكتروني: {str(e)})')

    context = {
        'order': order,
        'order_items': order_items,
    }
    return render(request, 'admin/order_detail.html', context)


@staff_member_required
def admin_categories(request):
    """Category management"""
    categories = Category.objects.all().order_by('name')
    return render(request, 'admin/categories.html', {'categories': categories})


@staff_member_required
def admin_category_add(request):
    """Add new category"""
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        image = request.FILES.get('image')

        try:
            category = Category.objects.create(
                name=name,
                description=description,
                image=image
            )
            messages.success(request, f'تم إضافة القسم "{category.name}" بنجاح')
            return redirect('admin_app:admin_categories')
        except Exception as e:
            messages.error(request, f'حدث خطأ: {str(e)}')

    return render(request, 'admin/category_add.html')


@staff_member_required
def admin_category_edit(request, category_id):
    """Edit category"""
    category = get_object_or_404(Category, id=category_id)

    if request.method == 'POST':
        category.name = request.POST.get('name', category.name)
        category.description = request.POST.get('description', category.description)

        if 'image' in request.FILES:
            category.image = request.FILES['image']

        category.save()
        messages.success(request, 'تم تحديث القسم بنجاح')
        return redirect('admin_app:admin_categories')

    return render(request, 'admin/category_edit.html', {'category': category})


@staff_member_required
def admin_category_delete(request, category_id):
    """Delete category"""
    if request.method == 'POST':
        category = get_object_or_404(Category, id=category_id)
        category_name = category.name
        category.delete()
        messages.success(request, f'تم حذف القسم {category_name}')
    return redirect('admin_app:admin_categories')


@staff_member_required
def admin_pending_users(request):
    """Display users awaiting approval"""
    from accounts.models import CustomUser
    pending_users = CustomUser.objects.filter(is_active=False, is_staff=False).order_by('-date_joined')
    context = {
        'pending_users': pending_users,
        'pending_users_active': 'active',
    }
    return render(request, 'admin/pending_users.html', context)


@staff_member_required
def admin_all_users(request):
    """Display all users"""
    from accounts.models import CustomUser
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')

    users = CustomUser.objects.filter(is_staff=False).order_by('-date_joined')

    if search_query:
        users = users.filter(
            Q(username__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(phone__icontains=search_query)
        )

    if status_filter == 'active':
        users = users.filter(is_active=True)
    elif status_filter == 'inactive':
        users = users.filter(is_active=False)

    context = {
        'users': users,
        'search_query': search_query,
        'status_filter': status_filter,
        'all_users_active': 'active',
    }
    return render(request, 'admin/all_users.html', context)


def _excel_safe_text(value):
    """Return untrusted text without allowing Excel formula injection."""
    if value is None:
        return ''
    value = str(value)
    if value.startswith(('=', '+', '-', '@')):
        return f"'{value}"
    return value


def _excel_datetime(value):
    """Convert an aware Django datetime to an Excel-compatible local datetime."""
    if value is None:
        return None
    if timezone.is_aware(value):
        value = timezone.localtime(value)
    return value.replace(tzinfo=None)


@staff_member_required
def admin_export_users_excel(request):
    """Export every user and their non-secret account data to an Arabic XLSX file."""
    from accounts.models import CustomUser, Profile
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    users = (
        CustomUser.objects.all()
        .select_related('profile')
        .prefetch_related('groups', 'user_permissions__content_type')
        .annotate(
            orders_count=Count('orders', distinct=True),
            support_messages_count=Count('customer_messages', distinct=True),
            cart_items_count=Count('cart__items', distinct=True),
            last_order_at=Max('orders__created_at'),
        )
        .order_by('id')
    )
    users_count = users.count()

    headers = [
        'رقم المستخدم',
        'اسم المستخدم',
        'الاسم الأول',
        'الاسم الأخير',
        'الاسم الكامل',
        'البريد الإلكتروني',
        'رقم الهاتف',
        'العنوان',
        'النبذة الشخصية',
        'صورة الملف الشخصي',
        'حالة الحساب',
        'نوع الحساب',
        'موظف',
        'مدير عام',
        'المجموعات',
        'الصلاحيات المباشرة',
        'تاريخ الانضمام',
        'آخر تسجيل دخول',
        'عدد الطلبات',
        'تاريخ آخر طلب',
        'عدد رسائل الدعم',
        'عدد عناصر السلة',
    ]

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'المستخدمون'
    worksheet.sheet_view.rightToLeft = True
    worksheet.freeze_panes = 'A5'
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.orientation = 'landscape'
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.print_title_rows = '1:4'

    workbook.properties.title = 'بيانات مستخدمي متجر الوسام'
    workbook.properties.subject = 'تصدير جميع المستخدمين من لوحة التحكم'
    workbook.properties.creator = 'متجر الوسام'

    last_column = worksheet.cell(row=1, column=len(headers)).column_letter
    worksheet.merge_cells(f'A1:{last_column}1')
    title_cell = worksheet['A1']
    title_cell.value = 'بيانات جميع مستخدمي متجر الوسام'
    title_cell.font = Font(name='Arial', size=16, bold=True, color='1A1A1A')
    title_cell.fill = PatternFill('solid', fgColor='FFC400')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    worksheet.row_dimensions[1].height = 30

    exported_at = timezone.localtime(timezone.now()).strftime('%Y/%m/%d %H:%M')
    worksheet.merge_cells(f'A2:{last_column}2')
    summary_cell = worksheet['A2']
    summary_cell.value = f'عدد المستخدمين: {users_count} — تاريخ التصدير: {exported_at}'
    summary_cell.font = Font(name='Arial', size=11, color='4B5563')
    summary_cell.alignment = Alignment(horizontal='right', vertical='center')
    worksheet.row_dimensions[2].height = 22

    header_fill = PatternFill('solid', fgColor='1A1A1A')
    header_font = Font(name='Arial', bold=True, color='FFFFFF')
    thin_gray = Side(style='thin', color='D1D5DB')
    cell_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    for column, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=4, column=column, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = cell_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    worksheet.row_dimensions[4].height = 32

    date_columns = {17, 18, 20}
    centered_columns = {1, 11, 12, 13, 14, 17, 18, 19, 20, 21, 22}

    for row_number, user in enumerate(users.iterator(chunk_size=500), start=5):
        try:
            profile = user.profile
        except Profile.DoesNotExist:
            profile = None

        image_url = ''
        if profile and profile.image:
            try:
                image_url = request.build_absolute_uri(profile.image.url)
            except ValueError:
                image_url = profile.image.name

        group_names = '، '.join(group.name for group in user.groups.all())
        permission_names = '، '.join(
            f'{permission.content_type.app_label}.{permission.codename}'
            for permission in user.user_permissions.all()
        )
        account_type = 'مدير عام' if user.is_superuser else ('موظف' if user.is_staff else 'عميل')

        values = [
            user.pk,
            _excel_safe_text(user.username),
            _excel_safe_text(user.first_name),
            _excel_safe_text(user.last_name),
            _excel_safe_text(user.get_full_name()),
            _excel_safe_text(user.email),
            _excel_safe_text(user.phone),
            _excel_safe_text(user.address),
            _excel_safe_text(profile.bio if profile else ''),
            _excel_safe_text(image_url),
            'نشط' if user.is_active else 'موقوف',
            account_type,
            'نعم' if user.is_staff else 'لا',
            'نعم' if user.is_superuser else 'لا',
            _excel_safe_text(group_names),
            _excel_safe_text(permission_names),
            _excel_datetime(user.date_joined),
            _excel_datetime(user.last_login),
            user.orders_count,
            _excel_datetime(user.last_order_at),
            user.support_messages_count,
            user.cart_items_count,
        ]

        for column, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row_number, column=column, value=value)
            cell.font = Font(name='Arial', size=10)
            cell.border = cell_border
            cell.alignment = Alignment(
                horizontal='center' if column in centered_columns else 'right',
                vertical='top',
                wrap_text=True,
            )
            if column in date_columns and value is not None:
                cell.number_format = 'yyyy/mm/dd hh:mm'
            if row_number % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='FFF9E6')

    data_last_row = max(4, worksheet.max_row)
    worksheet.auto_filter.ref = f'A4:{last_column}{data_last_row}'

    column_widths = [
        14, 22, 18, 18, 24, 32, 18, 40, 38, 45, 15,
        15, 12, 14, 28, 45, 21, 21, 14, 21, 18, 17,
    ]
    for index, width in enumerate(column_widths, start=1):
        worksheet.column_dimensions[worksheet.cell(row=4, column=index).column_letter].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"alwesam-users-{timezone.localdate().isoformat()}.xlsx"
    arabic_filename = f"مستخدمي-الوسام-{timezone.localdate().isoformat()}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = (
        f'attachment; filename="{filename}"; filename*=UTF-8\'\'{quote(arabic_filename)}'
    )
    response['Cache-Control'] = 'private, no-store'
    response['Pragma'] = 'no-cache'
    response['X-Content-Type-Options'] = 'nosniff'
    return response


@staff_member_required
def admin_approve_user(request, user_id):
    """Approve user registration"""
    from accounts.models import CustomUser
    if request.method == 'POST':
        user = get_object_or_404(CustomUser, id=user_id, is_staff=False)
        user.is_active = True
        user.save()

        try:
            from utils.email_tasks import send_activation_email_task
            login_url = request.build_absolute_uri('/accounts/login/')
            send_activation_email_task.delay(user.id, login_url)
            messages.success(request, f'تم تفعيل حساب "{user.username}" وتم إضافة إرسال البريد الإلكتروني إلى قائمة الانتظار')
        except Exception as e:
            messages.success(request, f'تم تفعيل حساب "{user.username}" (فشل جدولة إرسال البريد الإلكتروني: {str(e)})')

    return redirect('admin_app:pending_users')


@staff_member_required
def admin_reject_user(request, user_id):
    """Reject user and delete account"""
    from accounts.models import CustomUser
    if request.method == 'POST':
        user = get_object_or_404(CustomUser, id=user_id, is_staff=False)
        username = user.username
        user.delete()
        messages.success(request, f'تم رفض وحذف طلب انضمام "{username}"')
    return redirect('admin_app:pending_users')


@staff_member_required
def admin_toggle_user_status(request, user_id):
    """Toggle user active status (deactivate/activate)"""
    from accounts.models import CustomUser
    if request.method == 'POST':
        user = get_object_or_404(CustomUser, id=user_id, is_staff=False)
        user.is_active = not user.is_active
        user.save()

        if user.is_active:
            try:
                from utils.email_tasks import send_activation_email_task
                login_url = request.build_absolute_uri('/accounts/login/')
                send_activation_email_task.delay(user.id, login_url)
                messages.success(request, f'تم تفعيل حساب "{user.username}" وتم إرسال البريد الإلكتروني')
            except Exception as e:
                messages.success(request, f'تم تفعيل حساب "{user.username}" (فشل إرسال البريد الإلكتروني: {str(e)})')
        else:
            messages.success(request, f'تم إيقاف حساب المستخدم "{user.username}" بنجاح')

    return redirect('admin_app:all_users')
